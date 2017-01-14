import openpyxl
max=27 #AN
jk = openpyxl.load_workbook('Template.xlsx')
jk_sheet = jk.get_sheet_by_name('Sheet1')

for i in range(1,max+1):

	print "test 1-",i

	if (i<10):
		string='0'+str(i)
	else:
		string=str(i)
	xl_file_name="test"+string+".xlsx"

	district=openpyxl.load_workbook(xl_file_name)
	district_sheet=district.get_sheet_by_name('Sheet1')

	k=0

	while True:
		if (type(district_sheet.cell(row=4+1,column=k+2+1).value)==unicode):
			break
		k+=1
		print k

	

	
	#k adjusts for varied conversion from output file to test file 

	#district name
	jk_sheet.cell(row=i+1,column=0+1).value=district_sheet.cell(row=1+1,column=2+k+1).value

	#persons to SC % Vertically Down
	j=1
	a=4
	
	while (j<9):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=1+1+k+1).value 
		a=a+1
		j=j+1

	#nHouseholds
	jk_sheet.cell(row=i+1,column=9+1).value=district_sheet.cell(row=4+1,column=3+1+k+1).value
	#houseHoldSize
	jk_sheet.cell(row=i+1,column=10+1).value=district_sheet.cell(row=5+1,column=3+1+k+1).value
	#sexRation
	jk_sheet.cell(row=i+1,column=11+1).value=district_sheet.cell(row=7+1,column=3+1+k+1).value
	#sexRatio(0-6)
	jk_sheet.cell(row=i+1,column=12+1).value=district_sheet.cell(row=8+1,column=3+1+k+1).value
	#ST
	jk_sheet.cell(row=i+1,column=13+1).value=district_sheet.cell(row=10+1,column=3+1+k+1).value
	#%ST
	jk_sheet.cell(row=i+1,column=14+1).value=district_sheet.cell(row=11+1,column=3+1+k+1).value

	#literates
	jk_sheet.cell(row=i+1,column=15+1).value=district_sheet.cell(row=14+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=16+1).value=district_sheet.cell(row=15+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=17+1).value=district_sheet.cell(row=16+1,column=1+1+k+1).value

	#literacy_rate
	jk_sheet.cell(row=i+1,column=18+1).value=district_sheet.cell(row=18+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=19+1).value=district_sheet.cell(row=19+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=20+1).value=district_sheet.cell(row=20+1,column=1+1+k+1).value

	#education_whole
	j=21
	a=14
	while (j<28):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=3+1+k+1).value
		j=j+1
		a=a+1

	#age_groups
	j=28
	a=22
	while (j<32):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=3+1+k+1).value
		j=j+1
		a=a+1

	#workers
	j=32
	a=22
	while (j<36):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=1+1+k+1).value
		j=j+1
		a=a+1

	#sc_name
	jk_sheet.cell(row=i+1,column=36+1).value=district_sheet.cell(row=27+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=38+1).value=district_sheet.cell(row=28+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=40+1).value=district_sheet.cell(row=29+1,column=0+1+k+1).value

	#sc_pop
	jk_sheet.cell(row=i+1,column=37+1).value=district_sheet.cell(row=27+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=39+1).value=district_sheet.cell(row=28+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=41+1).value=district_sheet.cell(row=29+1,column=1+1+k+1).value

	#religeon_name
	jk_sheet.cell(row=i+1,column=42+1).value=district_sheet.cell(row=31+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=44+1).value=district_sheet.cell(row=32+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=46+1).value=district_sheet.cell(row=33+1,column=0+1+k+1).value

	#religeon_pop
	jk_sheet.cell(row=i+1,column=43+1).value=district_sheet.cell(row=31+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=45+1).value=district_sheet.cell(row=32+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=47+1).value=district_sheet.cell(row=33+1,column=1+1+k+1).value

	#st_name
	jk_sheet.cell(row=i+1,column=48+1).value=district_sheet.cell(row=27+1,column=2+1+k+1).value
	jk_sheet.cell(row=i+1,column=50+1).value=district_sheet.cell(row=28+1,column=2+1+k+1).value
	jk_sheet.cell(row=i+1,column=52+1).value=district_sheet.cell(row=29+1,column=2+1+k+1).value	

	#st_pop
	jk_sheet.cell(row=i+1,column=49+1).value=district_sheet.cell(row=27+1,column=3+1+k+1).value
	jk_sheet.cell(row=i+1,column=51+1).value=district_sheet.cell(row=28+1,column=3+1+k+1).value
	jk_sheet.cell(row=i+1,column=53+1).value=district_sheet.cell(row=29+1,column=3+1+k+1).value

	#imp_towns_names
	jk_sheet.cell(row=i+1,column=54+1).value=district_sheet.cell(row=39+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=56+1).value=district_sheet.cell(row=40+1,column=0+1+k+1).value
	jk_sheet.cell(row=i+1,column=58+1).value=district_sheet.cell(row=41+1,column=0+1+k+1).value

	#imp_towns_pop
	jk_sheet.cell(row=i+1,column=55+1).value=district_sheet.cell(row=39+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=57+1).value=district_sheet.cell(row=40+1,column=1+1+k+1).value
	jk_sheet.cell(row=i+1,column=59+1).value=district_sheet.cell(row=41+1,column=1+1+k+1).value

	#inhabited_villages
	jk_sheet.cell(row=i+1,column=60+1).value=district_sheet.cell(row=31+1,column=3+1+k+1).value

	#Ameneties
	j=61
	a=35
	while (j<77):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=3+1+k+1).value
		j=j+1
		a=a+1

	#houses
	j=77
	a=48
	while (j<80):
		jk_sheet.cell(row=i+1,column=j+1).value=district_sheet.cell(row=a+1,column=1+1+k+1).value
		j=j+1
		a=a+1

jk.save("final.xlsx")