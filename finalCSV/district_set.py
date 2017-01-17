"""
Removing crap from the district column
"""

import openpyxl
wb = openpyxl.load_workbook('final_all.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')

for i in range(2,sheet.max_row+1):
	print i
	district_name=str(sheet.cell(row=i,column=3).value).upper()
	test_district=0
	test_bracket=0

	if "DISTRICT" in district_name:
		test_district=1
	if "(" in district_name:
		test_bracket=1

	print "test_bracket",test_bracket
	print "test_district", test_district


	if (test_district==0 and test_bracket==0):
		continue

	elif (test_district==1 and test_bracket==0):
		sheet.cell(row=i,column=3).value=sheet.cell(row=i,column=3).value[9:]

	#finding the position of bracket

	j=0
	while True:
		if (district_name[j]=="("):
			break
		else:
			j+=1

	if (test_district==1 and test_bracket==1):
		sheet.cell(row=i,column=3).value=sheet.cell(row=i,column=3).value[9:j]

	elif(test_district==0 and test_bracket==1):
		sheet.cell(row=i,column=3).value=sheet.cell(row=i,column=3).value[0:j]


wb.save('final_all.xlsx')
