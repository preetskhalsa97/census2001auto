import openpyxl
from openpyxl.reader.excel import load_workbook

#EXTRACTING
wb = load_workbook('Extract.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')

for i in range(1,4):
	print (sheet['A'+str(i)].value)

print (sheet.cell(row=3, column=1).value)
a=sheet.cell(row=3, column=1).value
print a
b=sheet.cell(row=3, column=1)
print b.value

#PUTTING
book = load_workbook('Put.xlsx')
s= book.get_sheet_by_name('Sheet1')
s['A1']=4
s.cell(row=1,column=2).value=5
s.cell(row=1,column=3).value=b.value
s.cell(row=1,column=5).value=sheet.cell(row=3, column=1).value

for i in range(0,3):
	s.cell(row=i+1,column=5).value=sheet.cell(row=i, column=1).value







book.save('testPut.xlsx')















