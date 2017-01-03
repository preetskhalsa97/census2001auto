import openpyxl
max=14
jk = openpyxl.load_workbook('JK.xlsx')
jk_sheet = jk.get_sheet_by_name('Sheet1')
#print (jk_sheet.cell(row=0, column=1).value)

for i in range(1,max+1):
	if (i<10):
		string='0'+str(i)
	else:
		string=str(i)
	xl_file_name="output"+string+".xlsx"

	district=openpyxl.load_workbook(xl_file_name)
	district_sheet=district.get_sheet_by_name('Page 1')
	print (district_sheet.cell(row=0, column=7).value)
#############ISSUE###################### OUTPUT==> NONE############