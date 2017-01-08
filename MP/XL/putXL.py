import openpyxl
max=45 #MP
jk = openpyxl.load_workbook('Template.xlsx')
jk_sheet = jk.get_sheet_by_name('Sheet1')
jk.get_active_sheet()

for i in range(1,max+1):

	if (i==39 or i==44):
		continue

	if (i<10):
		string='0'+str(i)
	else:
		string=str(i)
	xl_file_name="test"+string+".xlsx"

	district=openpyxl.load_workbook(xl_file_name)
	district_sheet=district.get_sheet_by_name('Sheet1')

	k=0

	while True:
		if (type(district_sheet.cell(row=4,column=k+2).value)==int):
			break
		k+=1

	
	#k adjusts for varied conversion from output file to test file 

	#district name
	jk_sheet.cell(row=i,column=0).value=district_sheet.cell(row=1,column=2+k).value

	#persons to SC % Vertically Down
	j=1
	a=4
	
	while (j<9):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=1+1+k).value 
		a=a+1
		j=j+1

	#nHouseholds
	jk_sheet.cell(row=i,column=9).value=district_sheet.cell(row=4,column=3+1+k).value
	#houseHoldSize
	jk_sheet.cell(row=i,column=10).value=district_sheet.cell(row=5,column=3+1+k).value
	#sexRation
	jk_sheet.cell(row=i,column=11).value=district_sheet.cell(row=7,column=3+1+k).value
	#sexRatio(0-6)
	jk_sheet.cell(row=i,column=12).value=district_sheet.cell(row=8,column=3+1+k).value
	#ST
	jk_sheet.cell(row=i,column=13).value=district_sheet.cell(row=10,column=3+1+k).value
	#%ST
	jk_sheet.cell(row=i,column=14).value=district_sheet.cell(row=11,column=3+1+k).value

	#literates
	jk_sheet.cell(row=i,column=15).value=district_sheet.cell(row=14,column=1+1+k).value
	jk_sheet.cell(row=i,column=16).value=district_sheet.cell(row=15,column=1+1+k).value
	jk_sheet.cell(row=i,column=17).value=district_sheet.cell(row=16,column=1+1+k).value

	#literacy_rate
	jk_sheet.cell(row=i,column=18).value=district_sheet.cell(row=18,column=1+1+k).value
	jk_sheet.cell(row=i,column=19).value=district_sheet.cell(row=19,column=1+1+k).value
	jk_sheet.cell(row=i,column=20).value=district_sheet.cell(row=20,column=1+1+k).value

	#education_whole
	j=21
	a=14
	while (j<28):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=3+1+k).value
		j=j+1
		a=a+1

	#age_groups
	j=28
	a=22
	while (j<32):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=3+1+k).value
		j=j+1
		a=a+1

	#workers
	j=32
	a=22
	while (j<36):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=1+1+k).value
		j=j+1
		a=a+1

	#sc_name
	jk_sheet.cell(row=i,column=36).value=district_sheet.cell(row=27,column=0+1+k).value
	jk_sheet.cell(row=i,column=38).value=district_sheet.cell(row=28,column=0+1+k).value
	jk_sheet.cell(row=i,column=40).value=district_sheet.cell(row=29,column=0+1+k).value

	#sc_pop
	jk_sheet.cell(row=i,column=37).value=district_sheet.cell(row=27,column=1+1+k).value
	jk_sheet.cell(row=i,column=39).value=district_sheet.cell(row=28,column=1+1+k).value
	jk_sheet.cell(row=i,column=41).value=district_sheet.cell(row=29,column=1+1+k).value

	#religeon_name
	jk_sheet.cell(row=i,column=42).value=district_sheet.cell(row=31,column=0+1+k).value
	jk_sheet.cell(row=i,column=44).value=district_sheet.cell(row=32,column=0+1+k).value
	jk_sheet.cell(row=i,column=46).value=district_sheet.cell(row=33,column=0+1+k).value

	#religeon_pop
	jk_sheet.cell(row=i,column=43).value=district_sheet.cell(row=31,column=1+1+k).value
	jk_sheet.cell(row=i,column=45).value=district_sheet.cell(row=32,column=1+1+k).value
	jk_sheet.cell(row=i,column=47).value=district_sheet.cell(row=33,column=1+1+k).value

	#st_name
	jk_sheet.cell(row=i,column=48).value=district_sheet.cell(row=27,column=2+1+k).value
	jk_sheet.cell(row=i,column=50).value=district_sheet.cell(row=28,column=2+1+k).value
	jk_sheet.cell(row=i,column=52).value=district_sheet.cell(row=29,column=2+1+k).value	

	#st_pop
	jk_sheet.cell(row=i,column=49).value=district_sheet.cell(row=27,column=3+1+k).value
	jk_sheet.cell(row=i,column=51).value=district_sheet.cell(row=28,column=3+1+k).value
	jk_sheet.cell(row=i,column=53).value=district_sheet.cell(row=29,column=3+1+k).value

	#imp_towns_names
	jk_sheet.cell(row=i,column=54).value=district_sheet.cell(row=39,column=0+1+k).value
	jk_sheet.cell(row=i,column=56).value=district_sheet.cell(row=40,column=0+1+k).value
	jk_sheet.cell(row=i,column=58).value=district_sheet.cell(row=41,column=0+1+k).value

	#imp_towns_pop
	jk_sheet.cell(row=i,column=55).value=district_sheet.cell(row=39,column=1+1+k).value
	jk_sheet.cell(row=i,column=57).value=district_sheet.cell(row=40,column=1+1+k).value
	jk_sheet.cell(row=i,column=59).value=district_sheet.cell(row=41,column=1+1+k).value

	#inhabited_villages
	jk_sheet.cell(row=i,column=60).value=district_sheet.cell(row=31,column=3+1+k).value

	#Ameneties
	j=61
	a=35
	while (j<77):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=3+1+k).value
		j=j+1
		a=a+1

	#houses
	j=77
	a=48
	while (j<80):
		jk_sheet.cell(row=i,column=j).value=district_sheet.cell(row=a,column=1+1+k).value
		j=j+1
		a=a+1

jk.save("final.xlsx")